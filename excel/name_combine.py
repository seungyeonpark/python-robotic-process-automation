import openpyxl as excel

in_file = './input/name2.xlsx'
out_file = './output/name_combine.xlsx'

in_book = excel.load_workbook(in_file)
in_sheet = in_book.worksheets[0]

out_book = excel.Workbook()
out_sheet = out_book.active

for row in in_sheet.iter_rows():
    sung = row[0].value
    myung = row[1].value

    name = f"{sung} {myung}"

    out_sheet.append([name])

out_book.save(out_file)
