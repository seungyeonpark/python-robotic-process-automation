import openpyxl as excel
from openpyxl.styles import PatternFill
import datetime

in_file = './input/stock-data/xlsx'
out_file = './output/stock_highlight_oldtimes.xlsx'
limit = datetime.datetime(2020, 1, 1)


def check_row(row):
    date = row[0].value

    if type(date) is not datetime.datetime:
        return

    if date < limit:
        red = PatternFill(
            fill_type='solid',
            fgColor='ff0000'
        )

        for cell in row:
            cell.fill = red


def highlight_olditems(in_file_path, out_file_path):
    book = excel.load_workbook(in_file_path)

    for sheet in book.worksheets:
        for row in sheet.iter_rows(min_row=4):
            check_row(row)

    book.save(out_file)


if __name__ == '__main__':
    highlight_olditems(in_file, out_file)
