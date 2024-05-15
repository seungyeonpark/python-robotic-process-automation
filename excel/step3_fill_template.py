import openpyxl as excel
import json
import os
import datetime

template_file = './input/template/invoice-template.xlsx'
in_file = './output/split_data.json'

month = 3
out_dir = "./output/invoice_" + str(month)
subject = str(month) + "월 청구분"

if not os.path.exists(out_dir):
    os.mkdir(out_dir)

issue_date = datetime.datetime(2024, 5, 13).strftime('%Y/%m/%d')


def generate_invoice(name, data):
    book = excel.load_workbook(template_file)
    sheet = book.active

    sheet["G3"] = issue_date
    sheet["B4"] = name
    sheet["C10"] = subject
    sheet["C11"] = data["total"]

    for i, it in enumerate(data['items']):
        (date, summary, cnt, price) = it
        row = 15 + i

        sheet.cell(row, 2, f"{summary}({date})")
        sheet.cell(row, 5, cnt)
        sheet.cell(row, 6, price)
        sheet.cell(row, 7, cnt * price)

    out_file = f"{out_dir}/{name} 귀하.xlsx"
    book.save(out_file)
    print("save: ", out_file)


def fill_template():
    with open(in_file, "rt") as fp:
        users = json.load(fp)

    for name, data in users.items():
        generate_invoice(name, data)


if __name__ == "__main__":
    fill_template()