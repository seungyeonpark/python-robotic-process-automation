import openpyxl as excel
import glob

target_dir = './input/salesbook'
save_file = './output/merge_files.xlsx'


def merge_files():
    book = excel.Workbook()
    main_sheet = book.active
    enum_files(main_sheet)
    book.save(save_file)


def enum_files(main_sheet):
    files = glob.glob(target_dir + '/*.xlsx')
    for fname in files:
        read_book(main_sheet, fname)


def read_book(main_sheet, fname):
    print("read: ", fname)
    book = excel.load_workbook(fname, data_only=True)
    sheet = book.active

    for row in sheet.iter_rows(min_row=4):
        values = [cell.value for cell in row]

        if values[0] is None:
            break

        print(values)
        main_sheet.append(values)


if __name__ == "__main__":
    merge_files()

