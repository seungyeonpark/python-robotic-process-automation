import openpyxl as excel

in_file = './input/name1.xlsx'
out_file = './output/name_split.xlsx'

book = excel.load_workbook(in_file)
sheet = book.worksheets[0]

out_book = excel.Workbook()
out_sheet = out_book.active

for row in sheet.iter_rows():
    name = row[0].value.strip()

    if ' ' in name:
        (last_name, first_name) = name.split(' ')
    else:
        (last_name, first_name) = name[0], name[1:]

    out_sheet.append([last_name, first_name])

out_book.save(out_file)
