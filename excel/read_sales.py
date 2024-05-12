import openpyxl as excel

book = excel.load_workbook(
    "./input/monthly_sales.xlsx",
    data_only=True
)

sheet = book.active

# 1. 범위 명시적 지정
rows = sheet["A3":"F9"]

for row in rows:
    values = [cell.value for cell in row]
    print(values)

# 2. 최하단 행, 최하단 열
# 셀 테두리로 인해 실제 데이터보다 넓은 범위 반환할 수도 있음
print((sheet.max_row, sheet.max_column))

# 3. 모든 셀 가져오기
for row in sheet.iter_rows(min_row=3):
    values = [cell.value for cell in row]

    if values[0] is None:
        break

    print(values)

