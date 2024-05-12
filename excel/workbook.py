import openpyxl as excel

# 새 워크북 생성
book = excel.Workbook()

# 기존 워크북 열기
book = excel.load_workbook("./input/monthly_sales.xlsx")

# 워크북 열 때 수식을 계산하도록 설정
book = excel.load_workbook("./input/monthly_sales.xlsx", data_only=True)

# 명시적 종료
book.close()

# 활성 워크시트 가져오기
sheet = book.active

# 시트 번호로 불러오기
sheet = book.worksheets[0]

# 시트 이름으로 불러오기
sheet = book["Sheet1"]

# 문서에 포함된 시트 이름 목록 가져오기
print(book.sheetnames)

# 신규 시트 생성
sheet = book.create_sheet(title="new-sheet1")

# 기존 시트 복사
sheet = book.copy_worksheet(book["Sheet"])

# 시트 이름 변경
sheet.title = "new-sheet2"

# 시트 삭제
book.remove(book["Sheet1"])

