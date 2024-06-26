import openpyxl as excel
import datetime

book = excel.Workbook()
sheet = book.active

this_year = datetime.datetime.now().year

# header
sheet["A1"] = "출생 연도"
sheet["B2"] = "세는 나이"
sheet["C1"] = "만 나이 (생일 후)"
sheet["D1"] = "만 나이 (생일 전)"

sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 15

for i in range(80):
    birth_year = this_year - i
    korean_age = this_year - birth_year + 1
    man_age = {'after_bday' : korean_age - 1, 'before_bday' : korean_age - 2}

    year_cell = sheet.cell(i + 2, 1)
    year_cell.value = str(birth_year) + "년생"

    age_cell = sheet.cell(i + 2, 2)
    age_cell.value = str(korean_age) + "세"

    age_cell = sheet.cell(i + 2, 3)
    age_cell.value = "만" + str(man_age['after_bday']) + "세"

    age_cell = sheet.cell(i + 2, 4)
    age_cell.value = "만" + str(man_age['before_bday']) + "세"

sheet["D2"] = "-"

book.save("./output/write_agelist.xlsx")