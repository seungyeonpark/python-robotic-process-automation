import openpyxl as excel

template_file = './input/template/invoice-template.xlsx'
save_file = './output/invoice_simple.xlsx'

name = 'peter'
subject = '1월 청구분'
items = [
    ['사과', 5, 3200],
    ['바나나', 8, 2100],
    ['메론', 1, 12000]
]

book = excel.load_workbook(template_file)
sheet = book.active

sheet["B4"] = name
sheet["C10"] = subject

total = 0
for i, it in enumerate(items):
    (summary, count, price) = it
    subtotal = count * price
    total += subtotal

    row = 15 + i

    sheet.cell(row, 2, summary)
    sheet.cell(row, 5, count)
    sheet.cell(row, 6, price)
    sheet.cell(row, 7, subtotal)

sheet["C11"] = total

book.save(save_file)

