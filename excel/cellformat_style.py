import openpyxl as excel
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

book = excel.Workbook()
sheet = book.active

sheet.column_dimensions['B'].width = 40
sheet.row_dimensions[2].height = 40

cell = sheet["B2"]
cell.value = "text"

cell.alignment = Alignment(
    horizontal='center',
    vertical='center'
)

cell.border = Border(
    top=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
    left=Side(style='thin', color='000000')
)

cell.font = Font(
    size=14,
    bold=True,
    italic=True,
    color='FFFFFF'
)

cell.fill = PatternFill(
    fill_type='solid',
    fgColor='FF0000'
)

book.save("./output/cellformat_style.xlsx")

