import openpyxl as excel

book = excel.Workbook()
sheet = book.active

# header
sheet["A1"] = "출생 기간"
sheet["B1"] = "초등학교 입학 연도"
sheet["C1"] = "대학교 학번"

sheet.column_dimenstions['A'].width = 40
sheet.column_dimenstions['B'].width = 20
sheet.column_dimenstions['C'].width = 20

for i in range(50):
    birth_year = 2002 - i

    birth_range = "{}년 3월 1일생 ~ {}년 2월 28(29)일생".format(birth_year, birth_year + 1)

    ele_year = birth_year + 7

    univ_year = birth_year + 19
    univ_num = str(univ_year)[2:]

    sheet.cell(i + 2, 1, value=birth_range)
    sheet.cell(i + 2, 2, value=str(ele_year) + "년")
    sheet.cell(i + 2, 3, value=univ_num + "학번")

sheet["A2"] = "2002년 3월 1일생 ~ 2002년 12월 31일생"

book.save("./output/write_entry_year.xlsx")
