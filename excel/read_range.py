import openpyxl as excel

book = excel.load_workbook("./output/write_cellname.xlsx")
sheet = book.active

for y in range(2, 5):
    r = []

    for x in range(2, 5):
        v = sheet.cell(row=y, column=x).value
        r.append(v)

    print(r)

# sheet["좌측 상단 셀":"우측 하단 셀"]
# sheet["좌측 상단 셀:우측 하단 셀"]
for row in sheet["B2":"D4"]:
    r = []

    for cell in row:
        r.append(cell.value)

    print(r)

# iterator
it = sheet.iter_rows(
    min_row=2, min_col=2,
    max_row=4, max_col=4
)

for row in it:
    r = []

    for cell in row:
        r.append(cell.value)

    print(r)

