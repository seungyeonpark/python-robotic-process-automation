import openpyxl as excel

book = excel.Workbook()
sheet = book.active

# A1
sheet["A1"] = "value1"

# A2
sheet.cell(row=2, column=1, value="value2")

# A3
third_cell = sheet.cell(row=3, column=1)
third_cell.value = "value3"

book.save("./output/write_cell.xlsx")
