import openpyxl as excel
import csv

in_file = './input/name_addr.csv'
template_file = './input/template/card-template.xlsx'
save_file = './output/csv_card.xlsx'


def read_csv(fname):
    with open(fname, encoding='ansi') as f:
        reader = csv.reader(f)
        next(reader)
        return [v for v in reader]


book = excel.load_workbook(template_file)

sheet_template = book['Sheet']

for i, person in enumerate(read_csv(in_file)):
    (name, zip_num, addr) = person

    idx = i % 10

    if idx == 0:
        sheet = book.copy_worksheet(sheet_template)
        sheet.title = "Page" + str(i//10)

    row = 4 * (idx % 5) + 2
    col = 2 * (idx // 5) + 2

    sheet.cell(row=row+0, column=col, value=name)
    sheet.cell(row=row+1, column=col, value=zip_num)
    sheet.cell(row=row+2, column=col, value=addr)

book.remove(sheet_template)
book.save(save_file)
